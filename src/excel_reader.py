"""엑셀 파일을 읽어 플레이스홀더 데이터를 추출하는 모듈."""

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def read_excel(file_path: str | Path) -> list[dict[str, str]]:
    """엑셀 파일을 읽어 각 행을 딕셔너리 리스트로 반환한다.

    첫 번째 행은 헤더(플레이스홀더 이름)로 사용된다.

    Args:
        file_path: 엑셀 파일 경로

    Returns:
        [{"이름": "홍길동", "직급": "과장", ...}, ...]

    Raises:
        FileNotFoundError: 파일이 존재하지 않을 때
        ValueError: 헤더가 비어있거나 데이터 행이 없을 때
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError("엑셀 파일이 비어있습니다.")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    if not any(headers):
        raise ValueError("헤더 행이 비어있습니다.")

    data_rows = rows[1:]
    if not data_rows:
        raise ValueError("데이터 행이 없습니다.")

    result = []
    for row in data_rows:
        # 모든 셀이 None인 행은 건너뛴다
        if all(cell is None for cell in row):
            continue

        record = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            value = row[i] if i < len(row) else None
            record[header] = _format_value(value)
        result.append(record)

    return result


def _format_value(value: Any) -> str:
    """셀 값을 문자열로 변환한다."""
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value)
